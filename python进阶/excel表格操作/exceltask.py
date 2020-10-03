# import os
# cmd = r"D:\pypath\wget.exe https://github.com/jcyrss/baiyueheiyu/files/5250047/default.zip "
# os.system(cmd)
# print("下载完成")
import os
import re
import openpyxl
targetpath = r'D:\pypath\excel\data'
filedict = {}

#读取inf文件的内容，把信息以列表返回
def read_inf_file(filepath):
    with open (filepath,'r',encoding="gbk") as f :
        content = f.read()
    messagelist = []
    titlelist = []
    keyword = r'(.*?):(.*)'
    p = re.compile(keyword,re.MULTILINE)
    for match in p.finditer(content) :
        #表格第一行内容所需
        titlelist.append(match.group(1))
        messagelist.append(match.group(2))
    return messagelist,titlelist

#主程序，遍历每个inf文件，并且读取其内容，把信息写入excel表格
book = openpyxl.Workbook()
sh = book.active
sh.title = '基本信息'
sh.cell(1,1).value = '路径'
sh.cell(1,2).value = '文件名'
rows = 2    #从第二行开始往下加信息
for (dirpath,dirnames,filenames) in os.walk(targetpath) :
    for fn in filenames:
        if '.inf' in fn :
            filepath = os.path.join(dirpath,fn)
            filedict[fn] = dirpath
            #读取内容
            messagelist,titlelist = read_inf_file(filepath)

            #如果表格第一行内容没写，则写入
            if not sh.cell(1,3).value :
                cols1 = 3
                for value in titlelist:
                    sh.cell(1,cols1).value = value
                    cols1 +=1

            #写入对应的具体数据内容
            sh.cell(rows,1).value = dirpath     #写入路径
            sh.cell(rows,2).value = fn          #写入文件名
            cols = 3    #从第三列开始写数据
            for date in messagelist :
                sh.cell(rows,cols).value = date
                cols +=1
            rows +=1    #没写入一次inf文件数据，行数加一
book.save("读取信息.xlsx")